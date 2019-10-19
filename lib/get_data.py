from openpyxl import load_workbook
from openpyxl import Workbook

class Word():
    def __init__(self, jp = None, ch = None, ks = None, n = 0):
        self.jp = jp
        self.ch = ch
        self.ks = ks
        self.n  = n
    def get(self, index):
        """ 
            index: 
                0 -> jp 日文
                1 -> ch 中文
                2 -> ks 漢字
        """
        if index == 0:
            return self.jp
        elif index == 1:
            return self.ch
        elif index == 2:
            return self.ks

class XlwingsData():
    def __init__(self, file_name, sheet, name):
        None

class PyxlData():
    def __init__(self, file_name, sheet):
        wb = load_workbook(file_name)
        #ws = wb.active
        self.ws = wb.get_sheet_by_name(sheet)

        self.lessons = self.getLessons() # {lesson1:1,lesson2:5...}
        self.word_list_set = {}

    def getLessons(self):
        tmp = {}
        for i in range(0, self.ws.max_column):
            column = list(self.ws.columns)[i]
            if column[0].value != None and column[1].value == None:
                # use pyxl index:1~inf
                tmp[column[0].value]=i+1
        return tmp
    
    def getNextLesson(self, index):
        for i in range(index+1-1, self.ws.max_column):
            column = list(self.ws.columns)[i]
            if column[0].value != None and column[1].value == None:
                return i+1

    def getWords(self, lesson):
        if self.word_list_set.__contains__(lesson): # improve speed
            return self.word_list_set[lesson]
        
        col_lesson = self.lessons[lesson] # column number of the lesson
        next_lesson = self.getNextLesson(col_lesson)
        word_list = []
        
        for i in range(2, self.ws.max_row+1):
            # if the row is blank, break
            if self.ws.cell(row=i, column=col_lesson+1).value == None:
                break
            word = Word()

            for j in range(col_lesson+1, next_lesson):
                column_index = self.ws.cell(row=1, column=j).value
                v = self.ws.cell(row=i, column=j).value
                if column_index == "中文":
                    word.ch = v
                elif column_index == "平假名":
                    word.jp = v
                elif column_index == "漢字" or column_index == "漢字or英文":
                    word.ks = v
                elif column_index == "熟記程度":
                    word.n  = int(v)
            word_list.append(word)

        return word_list


class DataBase():
    def __init__(self, file_name, sheet):
        # if .file.xlsx exist
        if 1 == 2:# excel is opened
            # use xlwings
            self.xls = XlwingsData(file_name, sheet)
        else:
            # use pyxl
            self.xls = PyxlData(file_name, sheet)
    def getLessons(self):
        return self.xls.getLessons()

    def getWords(self, lesson):
        return self.xls.getWords(lesson)


if __name__ == "__main__":
    db = DataBase("../the_word.xlsx", "單字")
